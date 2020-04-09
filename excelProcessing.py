from openpyxl import load_workbook
from openpyxl import Workbook
# Import relevant modules fomr `openpyxl.utils`
from openpyxl.utils import get_column_letter, column_index_from_string
from textProcessing import to_handle_string, f15, handle1214, \
genre, handle_genre, genre15

def f():
    # open workbook
    wb = load_workbook('test1.xlsx')
    print(wb.get_sheet_names)

    # open sheet
    sheet = wb['Sheet1']
    print(sheet.title)

    anotherSheet = wb.active
    print(anotherSheet)

    # Retrieve the value of a certain cell
    print(sheet['A1'].value)

    #Select element 'B2' of your sheet
    c = sheet['B2']
    print(c.value)

    # Retrieve the row number of your element
    print(c.row)
    
    # Retrieve the column letter of your element
    print(c.column)

    # Retrieve the coordinates of the cell
    print(c.coordinate)

    # Retrieve cell value
    print(sheet.cell(row=1,column=2).value)

    # Print out values in column 2
    for i in range(1,4):
        print(i, sheet.cell(row=i, column=2).value)

    # Return 'A'
    print(get_column_letter(1))

    # Return '1'
    print(column_index_from_string('A'))

    # Retrieve the maximum amount of rows
    print(sheet.max_row)

    # Retrieve the maximum amount of columns
    print(sheet.max_column)

    # Print row per row
    for cell_object in sheet['A1':'C3']:
        for cell in cell_object:
            print(cell.coordinate, cell.value)
        print('--- END ---')

def f1():
    # D:\mySpreadsheets
   
    wb = load_workbook('D:\\mySpreadsheets\\books.xlsx')
    sheet = wb['Лист1']
    needed_cell = sheet['C53']
    #print(needed_cell.value)

    texts = []
    authors = []

    i = 0
    coordinates = [['C53','C55'],['D53','D55']]
    for i in range(0,2):
        for cell_object in sheet[coordinates[i][0]:coordinates[i][1]]:
            for cell in cell_object:
                if i == 0:
                    texts.append(cell.value)
                elif i == 1:
                    authors.append(cell.value)
        i += 1

    print(texts)
    print(authors)


    file = open('D:\\pyda\\прочитанное\\2020.txt','a',encoding='utf-8')
    #f = file.readlines()
    file.write('\n')
    i = 0
    for x in range(0,len(texts)):
        file.write(texts[i] + ' - ' + authors[i] + '\n')
        i += 1
    file.close()

def f():

    #wb = Workbook('D:\\mySpreadsheets\\books.xlsx')
    wb = load_workbook('D:\\mySpreadsheets\\books.xlsx')
    ws = wb.active
    #sheet = ws['Лист1']
    ws['C56'] = 'text'
    ws['D56'] = 'author'
    wb.save('D:\\mySpreadsheets\\books.xlsx')
    ws.cell(row=2)
def f2():

    
    texts = open('common_text.txt','r',encoding='utf-8')
    authors = open('common_author.txt','r',encoding='utf-8')
    texts_list = texts.readlines()
    authors_list = authors.readlines()
    wb = load_workbook('D:\\mySpreadsheets\\books.xlsx')
    ws = wb.active
    sheet = wb['Лист1']
    
    #print(ws.cell(row=292,column=3).value)
    i = 4
    for x in range(i,292):
        ws.cell(row=i,column=3).value = texts_list[i-4]
        i += 1

    j = 4
    for y in range(j,292):
        ws.cell(row=j,column=4).value = authors_list[j-4]
        j += 1
       
    texts.close()
    authors.close()
    wb.save('D:\\mySpreadsheets\\books.xlsx')

def f3():

    
    year_file = open('common_year.txt','r')
    years_list = year_file.readlines()

    j = 0
    for x in years_list:
        years_list[j] = years_list[j][:x.index('\n')] 
        j += 1

    wb = load_workbook('D:\\mySpreadsheets\\books.xlsx')
    sheet = wb['Лист1']

    i = 4
    for x in range(i,292):
        #sheet.cell(row=i,column=5).value = years_list[i-4][:years_list[i-4].index('\n')]
        sheet.cell(row=i,column=5).value = years_list[i-4]
        i += 1

    year_file.close()
    wb.save('D:\\mySpreadsheets\\books.xlsx')
def f15():
    
    file_address = '2015.txt'
    
    handled_file = open(file_address,'r',encoding='utf-8')
    file_list = handled_file.readlines()

    texts = []
    authors = []
    text = []
    texts2 = []
    authors2 = []
    years = []
    years2 = []

    i = 0
 
    for x in file_list:
        
        if x == '(303)\n': break
        if i >= 3:
            if '"' in x:
                text = to_handle_string(x)
                added_author = x[x.index(')')+1:x.index('"')].strip()
                for y in text:
                    texts2.append(y)
                    authors2.append(added_author)
                    years2.append('2015')
                authors.append(added_author)
                texts.append(text)
                years.append('2015')
                
                
        i += 1

    handled_file.close()

    wb = load_workbook('Books.xlsx')
    sheet = wb['Лист1']

    

    i = 4
    j = 3
    k = 0
    used_list = texts2
    
    for item in range(k,3):
        for x in range(i,len(used_list)+4):
            sheet.cell(row=i,column=j).value = used_list[i-4]
            i += 1
        k += 1
        j += 1
        i = 4
        if k == 1: used_list = authors2
        elif k == 2: used_list = years2
        
    

    wb.save('Books.xlsx')

def ff():

    data_list = handle1214()

    wb = load_workbook('Books.xlsx')
    sheet = wb['Лист1']

    k = 0
    q = 0
    i = 4
    j = 3
    used_list = data_list[k]
    for item in range(k,3):
        for x in range(i,len(used_list)+4):
            sheet.cell(row=i,column=j).value = used_list[i-4]
            i += 1
        k += 1
        if k == 3: break
        used_list = data_list[k]
        j += 1
        i = 4

    wb.save('Books.xlsx')

def fq():
   
    wb = load_workbook('Books.xlsx')
    sheet = wb['Лист1']
    sheet2 = wb['Лист3']
    texts = []
    authors = []

    i = 4
    col = 3
    for x in range(i,666):
        added_value = sheet.cell(row=i,column=col).value
        sheet2.cell(row=i,column=3).value = added_value
        print(added_value)
        texts.append(added_value)
        i += 1

    j = 4
    for x in range(j,666):
        added_value = sheet.cell(row=j,column=4).value
        sheet2.cell(row=j,column=4).value = added_value
        print(added_value)
        authors.append(added_value)
        j += 1

    #
    #k = 4
    #for x in range(k,666):
    #    sheet2.cell(row=k,column=3).value = texts[k-4]
    #    k += 1
    
    wb.save('Books.xlsx')

def fq1():

    genre_list = genre()
    
    wb = load_workbook('Books.xlsx')
    sheet = wb['Лист3']
    j = 0
    string = 378
    for x in range(string,666):
        sheet.cell(row=string,column=5).value = genre_list[string-378]
        string += 1
    wb.save('Books.xlsx')

def fq2():

    wb = load_workbook('Books.xlsx')
    sheet = wb['Лист3']
    i = 4
    for i in range(i,378):
        sheet.cell(row=i,column=5).value = 'художка'

    texts = genre15()
    string = 77
    j = 0
    for x in range(77,378):
        if (sheet.cell(row=string,column=3).value in texts):
            sheet.cell(row=string,column=5).value = 'нон-фикшн'
        
        string += 1
        
    wb.save('Books.xlsx')
        
        
               
       
    
        

    
    
        
            
    
    
