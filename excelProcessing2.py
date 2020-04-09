import openpyxl
from openpyxl import load_workbook
texts_nf = [[],[],[]]
    
def f():

    wb = load_workbook('books.xlsx')
    sheet1 = wb['Лист3']
    sheet2 = wb['Лист5']
    for x in range(4,711):
        if sheet1.cell(row=x, column=5).value == 'нон-фикшн':
            texts_nf[0].append(sheet1.cell(row=x, column=3).value)
            texts_nf[1].append(sheet1.cell(row=x, column=4).value)
            texts_nf[2].append(sheet1.cell(row=x, column=5).value)

    j = 3
    for x in range(3):
        k = 0
        for y in range(len(texts_nf[0])):
            sheet2.cell(row=k+4, column=j).value = texts_nf[j-3][k]
            k += 1
        j += 1

    wb.save('books.xlsx')
